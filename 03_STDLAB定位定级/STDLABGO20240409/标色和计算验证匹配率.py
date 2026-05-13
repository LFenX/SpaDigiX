import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = '澳门10-19零一验证数据-分类时间.xlsx'
xls = pd.ExcelFile(file_path)

# Initialize a dictionary to store the results
results = {
    '工作表名称': [],
    '总数据个数': [],
    '匹配的数据个数': [],
    '匹配率': []
}

# Load the workbook using openpyxl
wb = load_workbook(file_path)

# Process each sheet in the Excel file
for sheet_name in xls.sheet_names:
    df = pd.read_excel(xls, sheet_name=sheet_name)
    total_count = len(df)
    match_count = 0

    # Check if both columns exist in the sheet
    if '预测生肖段' in df.columns and '地支' in df.columns:
        # Iterate through each row to check for matches
        for index, row in df.iterrows():
            if str(row['地支']) in str(row['预测生肖段']):
                match_count += 1
                # Apply background color using openpyxl
                sheet = wb[sheet_name]
                fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                col_pred = df.columns.get_loc('预测生肖段') + 1
                col_dz = df.columns.get_loc('地支') + 1
                sheet.cell(row=index + 2, column=col_pred).fill = fill  # 预测生肖段 column
                sheet.cell(row=index + 2, column=col_dz).fill = fill  # 地支 column

    # Calculate the match ratio
    match_ratio = match_count / total_count if total_count > 0 else 0

    # Store the results
    results['工作表名称'].append(sheet_name)
    results['总数据个数'].append(total_count)
    results['匹配的数据个数'].append(match_count)
    results['匹配率'].append(match_ratio)

# Create a summary DataFrame
summary_df = pd.DataFrame(results)

# Add the summary sheet to the workbook
summary_sheet = wb.create_sheet('匹配率汇总')
for r_idx, row in enumerate(summary_df.itertuples(), 1):
    for c_idx, value in enumerate(row[1:], 1):
        summary_sheet.cell(row=r_idx + 1, column=c_idx, value=value)

# Save the modified Excel file
output_file_path = '含准确率-澳门10-19零一验证数据-分类时间.xlsx'
wb.save(output_file_path)

output_file_path
