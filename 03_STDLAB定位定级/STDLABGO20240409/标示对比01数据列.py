import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取Excel文件
file_path = '修正澳门-最低吻合度分段汇总数据结果.xlsx'
wb = load_workbook(file_path)
# 定义浅蓝色填充样式
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# 遍历每个工作表
for sheet in wb.sheetnames:
    ws = wb[sheet]

    # 遍历每一列，找到所有的“对比01数据”列
    for col in range(1, ws.max_column):
        col_letter = ws.cell(row=1, column=col).value
        if col_letter == "对比01数据":
            # 对比“对比01数据”列和其后面一列的数据
            for row in range(2, ws.max_row + 1):
                col1_value = ws.cell(row=row, column=col).value
                col2_value = ws.cell(row=row, column=col + 1).value
                if col1_value == col2_value:
                    ws.cell(row=row, column=col).fill = light_blue_fill
                    ws.cell(row=row, column=col + 1).fill = light_blue_fill

# 保存修改后的文件
output_path = '含标识-修正澳门-最低吻合度分段汇总数据结果.xlsx'
wb.save(output_path)