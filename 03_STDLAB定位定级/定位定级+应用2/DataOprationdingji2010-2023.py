import openpyxl
from openpyxl.styles import PatternFill
def caozuo(i):
    # 打开工作簿
    workbook = openpyxl.load_workbook(f'澳门G4MC_2010-2024_日干支分类_{i}_20240119.xlsx')
    # 遍历前十四个工作表
    for sheet_name in workbook.sheetnames[:10]:
        sheet = workbook[sheet_name]

        # 第12列的列号
        column_number = 13

        # 遍历每一行，检查第12列是否为0
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            print(f"\n处理第 {row_number} 行:")

            # 打印整行数据
            print(f"整行数据：{row}")

            # 打印第11列的值和数据类型
            column_value = row[column_number - 1]
            print(f"第11列的值：{column_value}")
            print(f"第11列的数据类型：{type(column_value)}")

            # 打印条件判断结果
            condition_result = column_value == 0
            print(f"条件判断结果：{condition_result}")

            # 如果满足条件，修改整行的底色为浅绿色
            if condition_result:
                print("满足条件，修改底色为浅蓝色")
                for cell in sheet[row_number]:
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    # 保存修改后的工作簿
    workbook.save(f'澳门三级G4MC_2010-2024_{i}240119.xlsx')
for i in range(0,6):
    caozuo(i)

