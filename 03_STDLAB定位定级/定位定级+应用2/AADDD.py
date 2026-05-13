import openpyxl
from openpyxl.styles import PatternFill

def color_cells(file_path, output_file_path):
    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Define the fill color for the cells (light blue)
    fill_color = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Loop through the cells in the '误差数串' column (assuming it's the 4th column)
    for row in range(2, sheet.max_row + 1):  # Starting from row 2 to skip the header
        cell = sheet.cell(row, 4)
        if cell.value == 1:
            cell.fill = fill_color

    # Save the modified workbook
    workbook.save(output_file_path)

# Replace 'your_excel_file.xlsx' with the path to your file
file_path = '连错规律(含干支历+双干+双干宫).xlsx'
output_file_path = '连错规律(含干支历+双干+双干宫).xlsx'

color_cells(file_path, output_file_path)
