import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import  math
# 读取Excel文件
file_path = '澳门data.xlsx'  # 替换为你的Excel文件路径
output_file_path = '处理后的数据.xlsx'  # 替换为你的输出文件路径

# 打开Excel文件
wb = openpyxl.load_workbook(file_path)
nianfen=int(input("请输入年份:"))
nianfenshenxiaodingwei=(nianfen-1995)%12
# 循环遍历所有工作表
for sheet_name in wb.sheetnames:
    # 读取当前工作表的数据
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_to_check = "生肖"  # 替换为你要判断的列名
    column_to_copy = "数字"  # 替换为你要复制数据的列名

    # 执行相同的操作，修改DataFrame的数据
    for index, row in df.iterrows():
        if pd.isnull(row[column_to_check]):
            dizhini=["亥","戌","酉","申","未","午","巳","辰","卯","寅","丑","子"]
            dizhinixiuzheng=[]
            for i in range(-nianfenshenxiaodingwei,-nianfenshenxiaodingwei+12):
                dizhinixiuzheng.append(dizhini[i])
            if not math.isnan(row[column_to_copy]):
                shengxiaodingewioperation_index = int(row[column_to_copy]) % 12
                df.at[index, column_to_check] = dizhinixiuzheng[shengxiaodingewioperation_index-1]

    date_column_name = "日期"
    df[date_column_name] = pd.to_datetime(df[date_column_name]).dt.strftime('%Y-%m-%d')
    columns_to_convert_to_int = ["序号", "数字"]  # 替换为你要转换为整数的列名
    for column in columns_to_convert_to_int:
        df[column] = df[column].apply(lambda x: str(int(x)) if pd.notna(x) and pd.notna(pd.to_numeric(x)) else x)
    df = df.iloc[:, 1:]

    # 保存修改后的数据到Excel文件
    with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        # 设置单元格样式
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.value = str(cell.value)  # 转换为字符串，确保不会出现异常
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # 自动换行
                cell.font = openpyxl.styles.Font(size=15)  # 设置字体大小
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'),
                                                     top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'))  # 添加边框

                # 设置每个单元格的宽度和高度
                worksheet.column_dimensions[cell.column_letter].width = 30  # 宽度
                worksheet.row_dimensions[cell.row].height = 60  # 高度
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.value = str(cell.value)  # 转换为字符串，确保不会出现异常
