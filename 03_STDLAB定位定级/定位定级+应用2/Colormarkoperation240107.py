import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from copy import copy
def caozuo(i):
    # 打开工作簿
    workbook = openpyxl.load_workbook(f'澳门MCG4修正2010-2024_Onlevel_日干支分类_{i}20240119.xlsx')

    # 遍历前十个工作表
    for sheet_name in workbook.sheetnames[:10]:
        sheet = workbook[sheet_name]

        # 将原始工作表转换为 DataFrame
        df = pd.DataFrame(sheet.values)
        # 设置标题
        df.columns = df.iloc[0]
        df = df[1:]

        # 调整列的顺序
        # 假设 "实际差值" 列的索引为 15，"定级得数" 列的索引为 13
        col_names = df.columns.tolist()
        actual_difference = col_names.pop(15)
        col_names.insert(2, actual_difference)  # 将 "实际差值" 插入 "定级得数" 后面

        # 重新排列 DataFrame 列
        df = df[col_names]

        # 清除原有工作表内容
        for row in sheet:
            for cell in row:
                cell.value = None

        # 将修改后的 DataFrame 写回工作表
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value

        # 对颜色填充等单元格样式操作
        for row in sheet.iter_rows(min_row=2):
            # 第 18 列的颜色填充
            if row[18].value == 0:  # 第 18 列的索引是 17
                for cell in row:
                    cell.fill = PatternFill(start_color="ADD8E6", fill_type="solid")

            # 第 3 列的范围判断和底色修改
            column_3_value = row[4].value
            if column_3_value:
                if 0 < column_3_value <= 1:
                    row[4].fill = PatternFill(start_color="90EE90", fill_type="solid")
                elif 1 < column_3_value <= 2:
                    row[4].fill = PatternFill(start_color="FFFF99", fill_type="solid")
                elif 2 < column_3_value <= 3:
                    row[4].fill = PatternFill(start_color="FFB6C1", fill_type="solid")
                elif 3 < column_3_value <= 4:
                    row[4].fill = PatternFill(start_color="D2B48C", fill_type="solid")

    # 保存修改后的工作簿
    workbook.save(f'澳门MC_2010-2024-0119G4_Onelevel_{i}.xlsx')
for i in range(0,6):
    caozuo(i)