from  F118X  import getSpaDigiXdingweideshu
from functionSpaDigiXAPPONE import getthebasicmessageofnineGrids
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment
# Function to apply center alignment to all cells in a worksheet
def apply_center_alignment_to_all_cells(ws):
    """Apply center alignment to all cells in an Excel worksheet."""
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
# Corrected function to perform the new calculations
def perform_new_calculations_corrected(df, time_dict):
    """ Perform new calculations with corrected date handling. """
    for index, row in df.iterrows():
        year = row['日期'].year
        month = row['日期'].month
        day = row['日期'].day
        for time_period, hour in time_dict.items():
            if time_period=="双支":
                for i in range(0,12):
                    ridizhi=getthebasicmessageofnineGrids(year,month,day,2*i)[1][3][1]
                    shichengdizhi=getthebasicmessageofnineGrids(year,month,day,2*i)[1][4][1]
                    if ridizhi==shichengdizhi:
                        hourr=2*i
                        break
                value=getSpaDigiXdingweideshu(year, month, day, hourr)
                df.at[index, time_period] = value[0]
                df.at[index, "1.1内外遁"] = value[1]
                df.at[index, "1.2驿马"] = value[2]
                df.at[index, "1.3同边要素"] = value[3]
                df.at[index, "1.4空亡"] = value[4]
                df.at[index, "1.5支冲"] = value[5]
                df.at[index,"1.6伏吟"]=value[6]
                df.at[index, "1.7反吟"] = value[7]
                df.at[index, "1.8甲子戊+吉门"] = value[8]
                df.at[index, "2吉凶格"] = value[9]
                df.at[index, "3天盘干"] = value[10]
                df.at[index, "4九星"] = value[11]
                df.at[index, "5八门"] = value[12]
                df.at[index, "6八神"] = value[13]
                df.at[index, "7长生十二宫"] = value[14]
                df.at[index, "8阴阳遁新增"] = value[15]

            elif time_period=="双干":
                for i in range(0,12):
                    riganzhi=getthebasicmessageofnineGrids(year,month,day,2*i)[1][3]
                    shichengganzhi = getthebasicmessageofnineGrids(year, month, day, 2 * i)[1][4]
                    if riganzhi=="甲子" and shichengganzhi=="甲子":
                        for j in range(0,12):
                            panduanwugan=getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                            if panduanwugan=="戊":
                                hourr=2*j
                                break
                    elif riganzhi == "甲戌" and shichengganzhi == "甲戌":
                        for j in range(0,12):
                            panduanwugan=getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                            if panduanwugan=="己":
                                hourr=2*j
                                break
                    elif riganzhi == "甲申" and shichengganzhi == "甲申":
                        for j in range(0,12):
                            panduanwugan=getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                            if panduanwugan=="庚":
                                hourr=2*j
                                break
                    elif riganzhi == "甲寅" and shichengganzhi == "甲子":
                        for j in range(0,12):
                            panduanwugan=getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                            if panduanwugan=="癸":
                                hourr=2*j
                                break
                    elif riganzhi == "甲午" and shichengganzhi == "甲午":
                        for j in range(0,12):
                            panduanwugan=getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                            if panduanwugan=="辛":
                                hourr=2*j
                                break
                    elif riganzhi == "甲辰" and shichengganzhi == "甲辰":
                        for j in range(0,12):
                            panduanwugan=getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                            if panduanwugan=="壬":
                                hourr=2*j
                                break
                    elif riganzhi[0] == "甲" and shichengganzhi[0] == "甲":
                        if riganzhi == "甲子":
                            for j in range(0, 12):
                                panduanwugan = getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                                if panduanwugan == "戊":
                                    shuanggan = "戊"
                                    hourr = 2 * j
                                    break
                        elif riganzhi == "甲戌":
                            for j in range(0, 12):
                                panduanwugan = getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                                if panduanwugan == "己":
                                    shuanggan = "己"
                                    hourr = 2 * j
                                    break
                        elif riganzhi == "甲辰":
                            for j in range(0, 12):
                                panduanwugan = getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                                if panduanwugan == "壬":
                                    shuanggan = "壬"
                                    hourr = 2 * j
                                    break
                        elif riganzhi == "甲寅":
                            for j in range(0, 12):
                                panduanwugan = getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                                if panduanwugan == "癸":
                                    shuanggan = "癸"
                                    hourr = 2 * j
                                    break
                        elif riganzhi == "甲申":
                            for j in range(0, 12):
                                panduanwugan = getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                                if panduanwugan == "庚":
                                    shuanggan = "庚"
                                    hourr = 2 * j
                                    break
                        elif riganzhi == "甲午":
                            for j in range(0, 12):
                                panduanwugan = getthebasicmessageofnineGrids(year, month, day, 2 * j)[1][4][0]
                                if panduanwugan == "辛":
                                    shuanggan = "辛"
                                    hourr = 2 * j
                                    break
                    else:
                        if riganzhi[0]==shichengganzhi[0]:
                            hourr=2*i
                            break


                value = getSpaDigiXdingweideshu(year, month, day, hourr)
                df.at[index, time_period] = value[0]
                df.at[index, "1.1内外遁"] = value[1]
                df.at[index, "1.2驿马"] = value[2]
                df.at[index, "1.3同边要素"] = value[3]
                df.at[index, "1.4空亡"] = value[4]
                df.at[index, "1.5支冲"] = value[5]
                df.at[index, "1.6伏吟"] = value[6]
                df.at[index, "1.7反吟"] = value[7]
                df.at[index, "1.8甲子戊+吉门"] = value[8]
                df.at[index, "2吉凶格"] = value[9]
                df.at[index, "3天盘干"] = value[10]
                df.at[index, "4九星"] = value[11]
                df.at[index, "5八门"] = value[12]
                df.at[index, "6八神"] = value[13]
                df.at[index, "7长生十二宫"] = value[14]
                df.at[index, "8阴阳遁新增"] = value[15]

    return df
# Time periods dictionary
time_dict = {"双支":0}

# Load the Excel file
file_path_source = '定位、定级验证灵数误差SDX118SZ表.xlsx'
sheet_names = ['澳门新定位结果', '香港新定位结果']

# Reading data from the sheets, including the error sheets
data_1_source = pd.read_excel(file_path_source, sheet_name=sheet_names[0], parse_dates=['日期'])
data_2_source = pd.read_excel(file_path_source, sheet_name=sheet_names[1], parse_dates=['日期'])
data_error_1_source = pd.read_excel(file_path_source, sheet_name='澳门定位误差', parse_dates=['日期'])
data_error_2_source = pd.read_excel(file_path_source, sheet_name='香港定位误差', parse_dates=['日期'])
data_error_1_source_abs=pd.read_excel(file_path_source, sheet_name='澳门定位误差ABS', parse_dates=['日期'])
data_error_2_source_abs=pd.read_excel(file_path_source, sheet_name='香港定位误差ABS', parse_dates=['日期'])
# Perform calculations on the source sheets
data_1_calculated_source = perform_new_calculations_corrected(data_1_source.copy(), time_dict)
data_2_calculated_source = perform_new_calculations_corrected(data_2_source.copy(), time_dict)

# Calculate errors and update the error sheets for both '澳门定位误差' and '香港定位误差'
for index, row in data_1_calculated_source.iterrows():
    for time_period in time_dict.keys():
        if time_period =="双支":
            error_value_1 = data_1_source.at[index, '双支灵数'] - row[time_period]
            data_error_1_source.at[index, time_period + '误差'] = error_value_1
        else:
            error_value_1 = data_1_source.at[index, '双干灵数'] - row[time_period]
            data_error_1_source.at[index, time_period + '误差'] = error_value_1



for index, row in data_2_calculated_source.iterrows():
    for time_period in time_dict.keys():
        if time_period =="双支":
            error_value_2 = data_2_source.at[index, '双支灵数'] - row[time_period]
            data_error_2_source.at[index, time_period + '误差'] = error_value_2
        else:
            error_value_2 = data_2_source.at[index, '双干灵数'] - row[time_period]
            data_error_2_source.at[index, time_period + '误差'] = error_value_2

for index, row in data_1_calculated_source.iterrows():
    for time_period in time_dict.keys():
        if time_period=="双支":
            error_value_1_abs = abs(data_1_source.at[index, '双支灵数'] - row[time_period])
            data_error_1_source_abs.at[index, time_period + '误差'] = error_value_1_abs
        else:
            error_value_1_abs = abs(data_1_source.at[index, '双干灵数'] - row[time_period])
            data_error_1_source_abs.at[index, time_period + '误差'] = error_value_1_abs

for index, row in data_2_calculated_source.iterrows():
    for time_period in time_dict.keys():
        if time_period == "双支":
            error_value_2_abs = abs(data_2_source.at[index, '双支灵数'] - row[time_period])
            data_error_2_source_abs.at[index, time_period + '误差'] = error_value_2_abs
        else:
            error_value_2_abs = abs(data_2_source.at[index, '双干灵数'] - row[time_period])
            data_error_2_source_abs.at[index, time_period + '误差'] = error_value_2_abs

# Formatting the date column in the calculated dataframes
data_1_calculated_source['日期'] = data_1_calculated_source['日期'].dt.strftime('%Y-%m-%d')
data_2_calculated_source['日期'] = data_2_calculated_source['日期'].dt.strftime('%Y-%m-%d')
data_error_1_source['日期'] = data_error_1_source['日期'].dt.strftime('%Y-%m-%d')
data_error_2_source['日期'] = data_error_2_source['日期'].dt.strftime('%Y-%m-%d')
data_error_1_source_abs['日期'] = data_error_1_source_abs['日期'].dt.strftime('%Y-%m-%d')
data_error_2_source_abs['日期'] = data_error_2_source_abs['日期'].dt.strftime('%Y-%m-%d')
# Saving the results back to a new Excel file with calculations and updated error sheets
output_file_path_updated_source = '规则细分数据双支1220.xlsx'
with pd.ExcelWriter(output_file_path_updated_source, engine='openpyxl') as writer:
    data_1_calculated_source.to_excel(writer, sheet_name=sheet_names[0], index=False)
    data_2_calculated_source.to_excel(writer, sheet_name=sheet_names[1], index=False)
    data_error_1_source.to_excel(writer, sheet_name='澳门定位误差', index=False)
    data_error_2_source.to_excel(writer, sheet_name='香港定位误差', index=False)
    data_error_1_source_abs.to_excel(writer, sheet_name='澳门定位误差ABS', index=False)
    data_error_2_source_abs.to_excel(writer, sheet_name='香港定位误差ABS', index=False)

    # Apply center alignment to all cells in all worksheets
    workbook = writer.book
    for worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
        apply_center_alignment_to_all_cells(worksheet)