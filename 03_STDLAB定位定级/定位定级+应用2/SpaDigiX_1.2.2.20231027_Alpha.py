import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import  math
from functionSpaDigiXAPPTWO  import getthefourmaxmessage
# 读取Excel文件
file_path = '澳门data.xlsx'  # 替换为你的Excel文件路径
# 创建一个空的DataFrame
df = pd.DataFrame()
# 指定要保存的Excel文件名
output_file_path = '澳门data(预测).xlsx'
# 将空的DataFrame写入Excel文件
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='2023年新澳门', index=False)
output_file_path = '澳门data(预测).xlsx'  # 替换为你的输出文件路径
# 打开Excel文件
wb = openpyxl.load_workbook(file_path)
# 循环遍历所有工作表
for sheet_name in wb.sheetnames:
    # 读取当前工作表的数据
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_to_check = "生肖"  # 替换为你要判断的列名
    column_to_copy = "数字"  # 替换为你要复制数据的列名
    columns_to_check = ['序号', '日期', '数字']
    df = df.iloc[:, 1:11]
    nianfen = df.iloc[0,1].year
    nianfenshenxiaodingwei = int((nianfen - 1995) % 12)
    # 迭代DataFrame并删除包含空值的行
    for index, row in df.iterrows():
        if any(pd.isnull(row[column]) for column in columns_to_check):
            df.drop(index, inplace=True)

    # 执行相同的操作，修改DataFrame的数据
    for index, row in df.iterrows():
        if pd.isnull(row[column_to_check]):
            dizhini=["亥","戌","酉","申","未","午","巳","辰","卯","寅","丑","子"]
            dizhinixiuzheng=[]
            for i in range(-nianfenshenxiaodingwei,-nianfenshenxiaodingwei+12):
                dizhinixiuzheng.append(dizhini[i])
            if not math.isnan(row[column_to_copy]):
                shengxiaodingewioperation_index = int(row[column_to_copy]) % 12
                df.at[index, column_to_check] = dizhinixiuzheng[shengxiaodingewioperation_index-1]
    print(df)
    '''
#以下输入预测数据
#-----------------------------------------------------------------------------------------------------------------------------------------
#15天连续预测

    for i in range(len(df.index)):
        shengxiaozidonghuoqu=""
        if 0 <= i < len(df.index) - 15:
            for j  in range(0,15):
                #生肖输入
                shengxiaozidonghuoqu=shengxiaozidonghuoqu+df.iloc[i+j,3]
            #起始日期年月日
            riqizidonghuoqu=df.iloc[i,1]
            yearrr=riqizidonghuoqu.year
            monthhh=riqizidonghuoqu.month
            dayyy=riqizidonghuoqu.day
            #查询日期年月日
            chaxunriqizidonghuoqu=df.iloc[i+15,1]
            yearrrchaxun=chaxunriqizidonghuoqu.year
            monthhhchaxun=chaxunriqizidonghuoqu.month
            dayyychaxun=chaxunriqizidonghuoqu.day
            shiwutianlianxuyucejieguoshuru=getthefourmaxmessage(yearrr,monthhh,dayyy,15,15,0,shengxiaozidonghuoqu,yearrrchaxun,monthhhchaxun,dayyychaxun)
            df.iloc[i+15,4]=shiwutianlianxuyucejieguoshuru
#15天连续预测
    print(df)
#147预测
    for i in range(len(df.index)):
        shengxiaozidonghuoqu = ""
        if 0 <= i < len(df.index) - 15:
            for j  in range(0,15):
                if (j+1)%3==1:
                    # 生肖输入
                    shengxiaozidonghuoqu = shengxiaozidonghuoqu+df.iloc[i+j,3]
            # 起始日期年月日
            riqizidonghuoqu = df.iloc[i, 1]
            yearrr = riqizidonghuoqu.year
            monthhh = riqizidonghuoqu.month
            dayyy = riqizidonghuoqu.day
            # 查询日期年月日
            chaxunriqizidonghuoqu = df.iloc[i + 15, 1]
            yearrrchaxun = chaxunriqizidonghuoqu.year
            monthhhchaxun = chaxunriqizidonghuoqu.month
            dayyychaxun = chaxunriqizidonghuoqu.day
            shiwutianlianxuyucejieguoshuru = getthefourmaxmessage(yearrr, monthhh, dayyy, 15, 1, 2,
                                                                  shengxiaozidonghuoqu, yearrrchaxun, monthhhchaxun,
                                                                  dayyychaxun)
            df.iloc[i + 15, 5] = shiwutianlianxuyucejieguoshuru
#147预测
    print(df)
#5天连续预测
    
    for i in range(len(df.index)):
        shengxiaozidonghuoqu=""
        if 0 <= i < len(df.index) - 5:
            for j  in range(0,5):
                # 生肖输入
                shengxiaozidonghuoqu = shengxiaozidonghuoqu + df.iloc[i + j, 3]
                # 起始日期年月日
            riqizidonghuoqu = df.iloc[i, 1]
            yearrr = riqizidonghuoqu.year
            monthhh = riqizidonghuoqu.month
            dayyy = riqizidonghuoqu.day
            # 查询日期年月日
            chaxunriqizidonghuoqu = df.iloc[i + 5, 1]
            yearrrchaxun = chaxunriqizidonghuoqu.year
            monthhhchaxun = chaxunriqizidonghuoqu.month
            dayyychaxun = chaxunriqizidonghuoqu.day
            shiwutianlianxuyucejieguoshuru = getthefourmaxmessage(yearrr, monthhh, dayyy, 5, 5, 0,
                                                                  shengxiaozidonghuoqu, yearrrchaxun, monthhhchaxun,
                                                                  dayyychaxun)
            df.iloc[i + 5, 6] = shiwutianlianxuyucejieguoshuru
#5天连续预测
    print(df)
#7天连续预测
    
    for i in range(len(df.index)):
        shengxiaozidonghuoqu=""
        if 0 <= i < len(df.index) - 7:
            for j  in range(0,7):
                # 生肖输入
                shengxiaozidonghuoqu = shengxiaozidonghuoqu + df.iloc[i + j, 3]
                # 起始日期年月日
            riqizidonghuoqu = df.iloc[i, 1]
            yearrr = riqizidonghuoqu.year
            monthhh = riqizidonghuoqu.month
            dayyy = riqizidonghuoqu.day
            # 查询日期年月日
            chaxunriqizidonghuoqu = df.iloc[i + 7, 1]
            yearrrchaxun = chaxunriqizidonghuoqu.year
            monthhhchaxun = chaxunriqizidonghuoqu.month
            dayyychaxun = chaxunriqizidonghuoqu.day
            shiwutianlianxuyucejieguoshuru = getthefourmaxmessage(yearrr, monthhh, dayyy, 7, 7, 0,
                                                                  shengxiaozidonghuoqu, yearrrchaxun, monthhhchaxun,
                                                                  dayyychaxun)
            df.iloc[i + 7, 7] = shiwutianlianxuyucejieguoshuru
#7天连续预测
    print(df)
#10天连续预测
    for i in range(len(df.index)):
        shengxiaozidonghuoqu=""
        if 0 <= i < len(df.index) - 10:
            for j  in range(0,10):
                # 生肖输入
                shengxiaozidonghuoqu = shengxiaozidonghuoqu + df.iloc[i + j, 3]
                # 起始日期年月日
            riqizidonghuoqu = df.iloc[i, 1]
            yearrr = riqizidonghuoqu.year
            monthhh = riqizidonghuoqu.month
            dayyy = riqizidonghuoqu.day
            # 查询日期年月日
            chaxunriqizidonghuoqu = df.iloc[i + 10, 1]
            yearrrchaxun = chaxunriqizidonghuoqu.year
            monthhhchaxun = chaxunriqizidonghuoqu.month
            dayyychaxun = chaxunriqizidonghuoqu.day
            shiwutianlianxuyucejieguoshuru = getthefourmaxmessage(yearrr, monthhh, dayyy, 10, 10, 0,
                                                                  shengxiaozidonghuoqu, yearrrchaxun, monthhhchaxun,
                                                                  dayyychaxun)
            df.iloc[i + 10, 8] = shiwutianlianxuyucejieguoshuru
#10天连续预测
    print(df)
    '''
#49天间隔七天预测
    for i in range(len(df.index)):
        shengxiaozidonghuoqu = ""
        if 0 <= i < len(df.index) - 49:
            for j  in range(0,49):
                if (j+1)%7==1:
                    # 生肖输入
                    shengxiaozidonghuoqu = shengxiaozidonghuoqu+df.iloc[i+j,3]
            # 起始日期年月日
            riqizidonghuoqu = df.iloc[i, 1]
            yearrr = riqizidonghuoqu.year
            monthhh = riqizidonghuoqu.month
            dayyy = riqizidonghuoqu.day
            # 查询日期年月日
            chaxunriqizidonghuoqu = df.iloc[i + 49, 1]
            yearrrchaxun = chaxunriqizidonghuoqu.year
            monthhhchaxun = chaxunriqizidonghuoqu.month
            dayyychaxun = chaxunriqizidonghuoqu.day
            shiwutianlianxuyucejieguoshuru = getthefourmaxmessage(yearrr, monthhh, dayyy, 49, 1, 6,
                                                                  shengxiaozidonghuoqu, yearrrchaxun, monthhhchaxun,
                                                                  dayyychaxun)
            df.iloc[i + 49,9] = shiwutianlianxuyucejieguoshuru
#49天间隔七天预测
    print(df)
#-----------------------------------------------------------------------------------------------------------------------------------------
#以上输入预测数据

    date_column_name = "日期"
    df[date_column_name] = pd.to_datetime(df[date_column_name]).dt.strftime('%Y-%m-%d')
    columns_to_convert_to_int = ["序号", "数字"]  # 替换为你要转换为整数的列名
    for column in columns_to_convert_to_int:
        df[column] = df[column].apply(lambda x: str(int(x)) if pd.notna(x) and pd.notna(pd.to_numeric(x)) else x)



    # 保存修改后的数据到Excel文件
    with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        # 设置单元格样式
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.value = str(cell.value)  # 转换为字符串，确保不会出现异常
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # 自动换行
                cell.font = openpyxl.styles.Font(size=20)  # 设置字体大小
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'),
                                                     top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'))  # 添加边框

                # 设置每个单元格的宽度和高度
                worksheet.column_dimensions[cell.column_letter].width = 70  # 宽度
                worksheet.row_dimensions[cell.row].height = 120  # 高度
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.value = str(cell.value)  # 转换为字符串，确保不会出现异常
